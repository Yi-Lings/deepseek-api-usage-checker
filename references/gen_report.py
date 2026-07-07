#!/usr/bin/env python3
"""
DeepSeek API Usage Report Generator

Reads three JSON files (from playwright-cli response-body) and produces
a formatted Excel report with 4 worksheets.

Usage:
  python gen_report.py <report-dir> <tmp-dir>

Expects <tmp-dir>/ds_11.json, ds_12.json, ds_13.json.
Outputs <report-dir>/ds-<timestamp>.xlsx.
"""

import json, os, sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

REPORT_DIR = sys.argv[1]
WIN_TMP = sys.argv[2]

HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HDR_FONT = Font(bold=True, size=11, color="FFFFFF")
DATA_FONT = Font(size=11)
BOLD_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=12)
CENTER = Alignment(horizontal="center")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def rd(fn):
    with open(fn, encoding="utf-8") as f:
        return json.load(f)


def style_hdr(ws, n):
    for ci in range(1, n + 1):
        c = ws.cell(row=1, column=ci)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = CENTER
        c.border = BORDER


def sc(cell, hdr=False):
    cell.font = BOLD_FONT if hdr else DATA_FONT
    cell.alignment = CENTER
    cell.border = BORDER


def cr(hit, miss):
    t = hit + miss
    return f"{hit / t * 100:.2f}%" if t else "N/A"


os.makedirs(REPORT_DIR, exist_ok=True)

d11 = rd(os.path.join(WIN_TMP, "ds_11.json"))
d12 = rd(os.path.join(WIN_TMP, "ds_12.json"))
d13 = rd(os.path.join(WIN_TMP, "ds_13.json"))

s = d11.get("data", {}).get("biz_data", {})
bn = float((s.get("normal_wallets") or [{}])[0].get("balance", 0))
bb = float((s.get("bonus_wallets") or [{}])[0].get("balance", 0))
mc = float((s.get("monthly_costs") or [{}])[0].get("amount", 0))
mt = int(s.get("monthly_token_usage", 0))
av = int(s.get("total_available_token_estimation", 0))

a12 = d12.get("data", {}).get("biz_data", {})
c13 = d13.get("data", {}).get("biz_data", [])

amt_totals = a12.get("total", [])
cost_totals = c13[0].get("total", []) if c13 else []
amt_days = a12.get("days", [])

models_amt = {}
for m in amt_totals:
    models_amt[m["model"]] = {x["type"]: int(x["amount"]) for x in m["usage"]}
models_cost = {}
for m in cost_totals:
    models_cost[m["model"]] = {x["type"]: float(x["amount"]) for x in m["usage"]}
names = list(models_amt.keys())

wb = Workbook()
now = datetime.now()

# Sheet 1: Account Overview
ws1 = wb.active
ws1.title = "账户总览"
ws1.merge_cells("A1:C1")
ws1["A1"].value = f"DeepSeek API 用量报告 — {now.year}年{now.month}月"
ws1["A1"].font = TITLE_FONT
ws1["A3"] = "账户总览"
ws1["A3"].font = SECTION_FONT
for i, (k, v, u) in enumerate(
    [
        ("充值余额 (CNY)", round(bn, 2), "CNY"),
        ("赠送余额 (CNY)", round(bb, 2), "CNY"),
        ("月消费 (CNY)", round(mc, 2), "CNY"),
        ("月 Token 总消耗", mt, "Tokens"),
        ("可用 Token 估值", av, "Tokens"),
    ],
    start=4,
):
    for ci, val in enumerate([k, v, u], 1):
        sc(ws1.cell(row=i, column=ci, value=val), ci == 1)
ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 18
ws1.column_dimensions["C"].width = 12

# Sheet 2: Model Usage
ws2 = wb.create_sheet("模型用量")
h2 = [
    "模型",
    "请求数",
    "缓存命中 (Prompt)",
    "缓存未命中 (Prompt)",
    "输出 Tokens",
    "总 Tokens",
    "缓存命中率",
]
for ci, h in enumerate(h2, 1):
    ws2.cell(row=1, column=ci, value=h)
style_hdr(ws2, len(h2))
for ri, nm in enumerate(names, 2):
    u = models_amt[nm]
    hit = u.get("PROMPT_CACHE_HIT_TOKEN", 0)
    miss = u.get("PROMPT_CACHE_MISS_TOKEN", 0)
    resp = u.get("RESPONSE_TOKEN", 0)
    req = u.get("REQUEST", 0)
    for ci, v in enumerate(
        [nm, req, hit, miss, resp, hit + miss + resp, cr(hit, miss)], 1
    ):
        sc(ws2.cell(row=ri, column=ci, value=v))
ws2.column_dimensions["A"].width = 30
for c in range(2, 8):
    ws2.column_dimensions[get_column_letter(c)].width = 20

# Sheet 3: Cost Breakdown
ws3 = wb.create_sheet("消费明细")
h3 = ["模型", "缓存命中 (¥)", "缓存未命中 (¥)", "输出 (¥)", "合计 (¥)"]
for ci, h in enumerate(h3, 1):
    ws3.cell(row=1, column=ci, value=h)
style_hdr(ws3, len(h3))
gh = gm = gr = 0.0
for ri, nm in enumerate(names, 2):
    u = models_cost.get(nm, {})
    h = u.get("PROMPT_CACHE_HIT_TOKEN", 0)
    m = u.get("PROMPT_CACHE_MISS_TOKEN", 0)
    r = u.get("RESPONSE_TOKEN", 0)
    gh += h
    gm += m
    gr += r
    for ci, v in enumerate([nm, h, m, r, round(h + m + r, 4)], 1):
        sc(ws3.cell(row=ri, column=ci, value=v))
ri_tot = len(names) + 2
for ci, v in enumerate(
    ["总计", round(gh, 2), round(gm, 2), round(gr, 2), round(gh + gm + gr, 2)], 1
):
    sc(ws3.cell(row=ri_tot, column=ci, value=v), hdr=True)
ws3.column_dimensions["A"].width = 30
for c in range(2, 6):
    ws3.column_dimensions[get_column_letter(c)].width = 16

# Sheet 4: Daily Breakdown
ws4 = wb.create_sheet("按日分布")
h4 = [
    "日期",
    "模型",
    "请求数",
    "Input 缓存命中",
    "Input 缓存未命中",
    "输出 Tokens",
    "总 Tokens",
    "消费 (¥)",
]
for ci, h in enumerate(h4, 1):
    ws4.cell(row=1, column=ci, value=h)
style_hdr(ws4, len(h4))
cost_days = {}
if c13:
    for de in c13[0].get("days", []):
        d = de["date"]
        cost_days[d] = {}
        for m in de["data"]:
            cost_days[d][m["model"]] = {
                x["type"]: float(x["amount"]) for x in m["usage"]
            }
ri = 2
for de in amt_days:
    d = de["date"]
    for m in de["data"]:
        u = {x["type"]: int(x["amount"]) for x in m["usage"]}
        req = u.get("REQUEST", 0)
        if req == 0:
            continue
        hit = u.get("PROMPT_CACHE_HIT_TOKEN", 0)
        miss = u.get("PROMPT_CACHE_MISS_TOKEN", 0)
        resp = u.get("RESPONSE_TOKEN", 0)
        dc = cost_days.get(d, {}).get(m["model"], {})
        cv = sum(
            dc.get(t, 0)
            for t in [
                "PROMPT_CACHE_HIT_TOKEN",
                "PROMPT_CACHE_MISS_TOKEN",
                "RESPONSE_TOKEN",
            ]
        )
        for ci, v in enumerate(
            [d, m["model"], req, hit, miss, resp, hit + miss + resp, round(cv, 4)], 1
        ):
            sc(ws4.cell(row=ri, column=ci, value=v))
        ri += 1
ws4.column_dimensions["A"].width = 16
ws4.column_dimensions["B"].width = 24
for c in range(3, 9):
    ws4.column_dimensions[get_column_letter(c)].width = 18

fname = f"ds-{now.strftime('%y-%m-%d-%H-%M')}.xlsx"
fpath = os.path.join(REPORT_DIR, fname)
wb.save(fpath)
print(f"[OK] {fpath}")
